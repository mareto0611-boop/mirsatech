import os
import json
import sqlite3
import shutil
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, request, jsonify, session, send_file, render_template, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

# =========================
# App
# =========================

APP = Flask(__name__, template_folder="templates", static_folder="static")
APP.secret_key = os.environ.get("SECRET_KEY", "mirsa-tech-hospital-secret-dev")
APP.config["SESSION_PERMANENT"] = False

DEFAULT_PORT = 5000

ROLES = {
    "emergencias",
    "especialista",
    "planta",
    "enfermeria",
    "farmacia",
    "facturacion",
    "admision",
    "auditor",
    "administrador",
    "superadmin",
}

# =========================
# Config + Paths
# =========================

PROJECT_DIR = os.path.abspath(os.path.dirname(__file__))
LOCAL_DATA_DIR = os.path.join(PROJECT_DIR, "data")
os.makedirs(LOCAL_DATA_DIR, exist_ok=True)

CONFIG_PATH = os.path.join(LOCAL_DATA_DIR, "config.json")

def default_data_dir():
    appdata = os.getenv("APPDATA") or os.path.expanduser("~")
    return os.path.join(appdata, "MirsaTechHospital")

def load_config():
    cfg = {"data_dir": "", "port": DEFAULT_PORT}
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                file_cfg = json.load(f)
            if isinstance(file_cfg, dict):
                cfg.update(file_cfg)
        except Exception:
            pass

    data_dir = (cfg.get("data_dir") or "").strip()
    if not data_dir:
        data_dir = default_data_dir()
    port = cfg.get("port") or DEFAULT_PORT

    os.makedirs(data_dir, exist_ok=True)
    return {"data_dir": data_dir, "port": int(port)}

CONFIG = load_config()

DB_DIR = CONFIG["data_dir"]
DB_PATH = os.path.join(DB_DIR, "hospital.db")

UPLOADS_DIR = os.path.join(DB_DIR, "uploads")
BRANDING_DIR = os.path.join(UPLOADS_DIR, "branding")
os.makedirs(BRANDING_DIR, exist_ok=True)

EXPORTS_DIR = os.path.join(DB_DIR, "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

BACKUPS_DIR = os.path.join(DB_DIR, "backups")
os.makedirs(BACKUPS_DIR, exist_ok=True)

# =========================
# DB helpers
# =========================

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn

def now_local():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def get_setting(key, default=""):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = cur.fetchone()
    conn.close()
    if row and row["value"] is not None and str(row["value"]).strip() != "":
        return row["value"]
    return default

def set_setting(key, value):
    conn = db()
    cur = conn.cursor()
    cur.execute("INSERT OR REPLACE INTO settings(key, value) VALUES (?,?)", (key, str(value)))
    conn.commit()
    conn.close()

def audit(action, entity=None, entity_id=None, before=None, after=None):
    conn = db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO audit_log(created_at, user, role, action, entity, entity_id, before_json, after_json)
           VALUES (?,?,?,?,?,?,?,?)""",
        (
            now_local(),
            session.get("user", {}).get("username"),
            session.get("user", {}).get("role"),
            action,
            entity,
            entity_id,
            json.dumps(before, ensure_ascii=False) if before is not None else None,
            json.dumps(after, ensure_ascii=False) if after is not None else None,
        ),
    )
    conn.commit()
    conn.close()

def init_db():
    conn = db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      username TEXT UNIQUE NOT NULL,
      name TEXT,
      password_hash TEXT NOT NULL,
      role TEXT NOT NULL,
      active INTEGER NOT NULL DEFAULT 1,
      must_change_password INTEGER NOT NULL DEFAULT 0,
      created_at TEXT NOT NULL
    )
    """)

    # migrate: exequatur
    try:
        cur.execute("ALTER TABLE users ADD COLUMN exequatur TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS settings(
      key TEXT PRIMARY KEY,
      value TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS pacientes(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      created_at TEXT NOT NULL,
      paciente TEXT NOT NULL,
      diagnostico TEXT NOT NULL,
      tratamiento TEXT NOT NULL,
      tipo TEXT NOT NULL DEFAULT 'emergencias',
      estado TEXT NOT NULL DEFAULT 'pendiente',
      medico_user TEXT,
      enfermera_user TEXT,
      aplicado_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS materiales_catalogo(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      nombre TEXT UNIQUE NOT NULL,
      activo INTEGER NOT NULL DEFAULT 1,
      created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS materiales_paciente(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      paciente_id INTEGER NOT NULL,
      material_id INTEGER NOT NULL,
      cantidad INTEGER NOT NULL DEFAULT 1,
      updated_at TEXT NOT NULL,
      updated_by TEXT,
      UNIQUE(paciente_id, material_id),
      FOREIGN KEY(paciente_id) REFERENCES pacientes(id) ON DELETE CASCADE,
      FOREIGN KEY(material_id) REFERENCES materiales_catalogo(id) ON DELETE RESTRICT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS audit_log(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      created_at TEXT NOT NULL,
      user TEXT,
      role TEXT,
      action TEXT NOT NULL,
      entity TEXT,
      entity_id INTEGER,
      before_json TEXT,
      after_json TEXT
    )
    """)
    # -------------------------
    # NUEVO: Ingresos + Admisión + Evoluciones + Órdenes + Epicrisis (Egreso)
    # -------------------------

    cur.execute("""
    CREATE TABLE IF NOT EXISTS ingresos(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,

      estado TEXT NOT NULL DEFAULT 'activo', -- activo | egresado

      creado_por_user TEXT NOT NULL,
      creado_por_role TEXT NOT NULL,

      -- especialista que ordena (opcional)
      especialista_user TEXT,

      -- ubicación
      sala TEXT,
      habitacion TEXT,
      cama TEXT,

      -- info general
      especialidad TEXT,
      fecha_ingreso TEXT,
      hora_ingreso TEXT
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ingresos_estado ON ingresos(estado)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ingresos_created ON ingresos(created_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ingresos_especialista ON ingresos(especialista_user)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS ingresos_admision(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      ingreso_id INTEGER NOT NULL UNIQUE,

      updated_at TEXT NOT NULL,
      updated_by_user TEXT,
      updated_by_role TEXT,

      -- vínculo opcional con pacientes_master
      paciente_master_id INTEGER,

      -- Hoja de admisión (imagen 3)
      codigo TEXT,
      fecha TEXT,
      hora TEXT,
      habitacion TEXT,
      especialidad TEXT,

      nombres TEXT,
      apellidos TEXT,
      edad TEXT,
      sexo TEXT,
      direccion TEXT,
      cedula TEXT,
      seguro_ars TEXT,
      nss TEXT,
      ocupacion TEXT,

      fam_conyugue_acompanante TEXT,
      fam_madre TEXT,
      fam_padre TEXT,
      fam_direccion TEXT,
      fam_telefono TEXT,

      dx_1 TEXT,
      dx_2 TEXT,
      dx_3 TEXT,
      dx_4 TEXT,
      dx_5 TEXT,

      FOREIGN KEY(ingreso_id) REFERENCES ingresos(id) ON DELETE CASCADE,
      FOREIGN KEY(paciente_master_id) REFERENCES pacientes_master(id) ON DELETE SET NULL
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS ingresos_historia_clinica(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      ingreso_id INTEGER NOT NULL UNIQUE,

      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      created_by_user TEXT,
      created_by_role TEXT,
      updated_by_user TEXT,
      updated_by_role TEXT,

      fecha TEXT,
      hora TEXT,
      especialidad TEXT,

      nombres TEXT,
      apellidos TEXT,
      edad TEXT,
      direccion TEXT,

      motivo_consulta TEXT,
      historia_enfermedad_actual TEXT,

      app_1 TEXT,
      app_2 TEXT,
      app_3 TEXT,

      medicamentos_1 TEXT,
      medicamentos_2 TEXT,
      medicamentos_3 TEXT,

      aqx_1 TEXT,
      aqx_2 TEXT,
      aqx_3 TEXT,

      apnp_1 TEXT,
      apnp_2 TEXT,
      apnp_3 TEXT,

      alergicos TEXT,

      obst_g TEXT,
      obst_p TEXT,
      obst_a TEXT,
      obst_c TEXT,
      obst_fum TEXT,

      hab_alcohol INTEGER,
      hab_cafe INTEGER,
      hab_cafe_tazas TEXT,
      hab_tabaco INTEGER,
      hab_vaper INTEGER,
      hab_medicamentos INTEGER,
      hab_medicamentos_cual TEXT,

      sv_ta TEXT,
      sv_fc TEXT,
      sv_fcf TEXT,
      sv_fr TEXT,
      sv_glicemia TEXT,
      sv_temp TEXT,
      sv_sao2 TEXT,
      sv_peso TEXT,

      glasgow_o TEXT,
      glasgow_v TEXT,
      glasgow_m TEXT,
      glasgow_total TEXT,

      insp_aspecto_general TEXT,
      insp_cabeza TEXT,
      insp_cuello TEXT,
      insp_torax TEXT,
      insp_pulmones TEXT,
      insp_corazon TEXT,
      insp_abdomen TEXT,
      insp_genitales TEXT,
      insp_extremidades TEXT,
      insp_tacto_rectal TEXT,
      insp_neurologico TEXT,

      reportes TEXT,
      laboratorio TEXT,

      img_ekg TEXT,
      img_radiografia TEXT,
      img_sonografia TEXT,
      img_tomografia TEXT,

      comentarios TEXT,

      diag_1 TEXT,
      diag_2 TEXT,
      diag_3 TEXT,
      diag_4 TEXT,
      diag_5 TEXT,

      medicos TEXT,

      FOREIGN KEY(ingreso_id) REFERENCES ingresos(id) ON DELETE CASCADE
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_historia_ingreso ON ingresos_historia_clinica(ingreso_id)")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS ingreso_evoluciones(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      ingreso_id INTEGER NOT NULL,

      fecha_hora TEXT NOT NULL, -- editable por auditoría/médicos
      nota TEXT NOT NULL,

      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      created_by_user TEXT,
      created_by_role TEXT,

      FOREIGN KEY(ingreso_id) REFERENCES ingresos(id) ON DELETE CASCADE
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_evo_ingreso ON ingreso_evoluciones(ingreso_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_evo_fecha ON ingreso_evoluciones(fecha_hora)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS ingreso_ordenes(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      ingreso_id INTEGER NOT NULL,

      fecha_hora TEXT NOT NULL, -- editable
      medidas_generales TEXT NOT NULL,
      ordenes TEXT NOT NULL,

      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      created_by_user TEXT,
      created_by_role TEXT,

      FOREIGN KEY(ingreso_id) REFERENCES ingresos(id) ON DELETE CASCADE
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ord_ingreso ON ingreso_ordenes(ingreso_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ord_fecha ON ingreso_ordenes(fecha_hora)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS ingreso_epicrisis(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      ingreso_id INTEGER NOT NULL UNIQUE,

      fecha_egreso TEXT NOT NULL,
      destino TEXT NOT NULL, -- alta_medica | referido | fallecido | fuga | alta_peticion
      referido_hospital TEXT,

      -- Diagnósticos (enumerados) y otros campos del formulario de epicrisis
      dx_ingreso TEXT,
      dx_egreso TEXT,
      procedimientos TEXT,
      hallazgos TEXT,
      tratamiento TEXT,
      plan_seguimiento TEXT,
      causas_fallecimiento TEXT,

      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      created_by_user TEXT,
      created_by_role TEXT,

      FOREIGN KEY(ingreso_id) REFERENCES ingresos(id) ON DELETE CASCADE
    )
    """)
    # -------------------------
    # NUEVO: pacientes master (buscador)
    # -------------------------
    cur.execute("""
    CREATE TABLE IF NOT EXISTS pacientes_master(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      nombre TEXT NOT NULL,
      apodo TEXT,
      sexo TEXT,
      edad TEXT,
      fecha_nacimiento TEXT,
      cedula_pasaporte_nui TEXT,
      nss TEXT,
      aseguradora TEXT,
      grupo_sanguineo TEXT,
      alergico TEXT,
      direccion TEXT,
      nacionalidad TEXT,
      telefono TEXT
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pac_master_nombre ON pacientes_master(nombre)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pac_master_ced ON pacientes_master(cedula_pasaporte_nui)")

    # -------------------------
    # NUEVO: Historia clínica asegurados (2 páginas)
    # -------------------------
    cur.execute("""
    CREATE TABLE IF NOT EXISTS hc_asegurados(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      estado TEXT NOT NULL DEFAULT 'pendiente_medico',
      fact_user TEXT,
      medico_user TEXT,

      expediente_clinico TEXT,
      fecha TEXT,
      hora_llegada TEXT,
      triaje_prioridad TEXT,

      nombre TEXT,
      apodo TEXT,
      sexo TEXT,
      edad TEXT,
      fecha_nacimiento TEXT,
      cedula_pasaporte_nui TEXT,
      aseguradora TEXT,
      nss TEXT,
      grupo_sanguineo TEXT,
      alergico TEXT,
      direccion TEXT,
      nacionalidad TEXT,
      telefono TEXT,

      via_llegada TEXT,
      ambulancia_no TEXT,
      paramedico TEXT,
      acompanante TEXT,
      parentesco TEXT,
      acompanante_telefono TEXT,
      acompanante_direccion TEXT,

      motivo_cefalea INTEGER DEFAULT 0,
      motivo_tos INTEGER DEFAULT 0,
      motivo_palpitaciones INTEGER DEFAULT 0,
      motivo_hematemesis INTEGER DEFAULT 0,
      motivo_dolor_lumbar INTEGER DEFAULT 0,
      motivo_diarrea INTEGER DEFAULT 0,
      motivo_mareos INTEGER DEFAULT 0,
      motivo_epistaxis INTEGER DEFAULT 0,
      motivo_nauseas INTEGER DEFAULT 0,
      motivo_dolor_toracico INTEGER DEFAULT 0,
      motivo_hematuria INTEGER DEFAULT 0,
      motivo_herida_arma INTEGER DEFAULT 0,
      motivo_fiebre INTEGER DEFAULT 0,
      motivo_disnea INTEGER DEFAULT 0,
      motivo_vomitos INTEGER DEFAULT 0,
      motivo_dolor_abdominal INTEGER DEFAULT 0,
      motivo_otro TEXT,

      historia_enfermedad_actual TEXT,
      antecedentes_1 TEXT,
      antecedentes_2 TEXT,
      antecedentes_3 TEXT,

      obst_g TEXT,
      obst_p TEXT,
      obst_a TEXT,
      obst_c TEXT,
      fum TEXT,

      origen_enfermedad TEXT,
      origen_otro TEXT,

      ta TEXT,
      fc TEXT,
      fr TEXT,
      glicemia TEXT,
      sao2 TEXT,
      temp TEXT,
      peso TEXT,
      glasgow_o TEXT,
      glasgow_v TEXT,
      glasgow_m TEXT,
      glasgow_total TEXT,

      examen_aspecto_general TEXT,
      examen_cabeza TEXT,
      examen_cuello TEXT,
      examen_torax TEXT,
      examen_pulmones TEXT,
      examen_corazon TEXT,
      examen_tacto_rectal TEXT,
      examen_genitales TEXT,
      examen_abdomen TEXT,
      examen_extremidades TEXT,
      examen_neurologico TEXT,

      lab_hemograma INTEGER DEFAULT 0,
      lab_glicemia INTEGER DEFAULT 0,
      lab_urea INTEGER DEFAULT 0,
      lab_creatinina INTEGER DEFAULT 0,
      lab_ex_orina INTEGER DEFAULT 0,
      lab_tgo INTEGER DEFAULT 0,
      lab_tgp INTEGER DEFAULT 0,
      lab_troponinas INTEGER DEFAULT 0,
      lab_ck INTEGER DEFAULT 0,
      lab_cpk_mb INTEGER DEFAULT 0,
      lab_tp INTEGER DEFAULT 0,
      lab_tpt INTEGER DEFAULT 0,
      lab_inr INTEGER DEFAULT 0,
      lab_gases_arteriales INTEGER DEFAULT 0,
      lab_sodio INTEGER DEFAULT 0,
      lab_potasio INTEGER DEFAULT 0,
      lab_cloro INTEGER DEFAULT 0,
      lab_coprologico INTEGER DEFAULT 0,
      lab_otros TEXT,

      lab_via_ind_ambulatoria INTEGER DEFAULT 0,
      lab_via_ind_emergencias INTEGER DEFAULT 0,

      img_ekg INTEGER DEFAULT 0,
      img_rayosx INTEGER DEFAULT 0,
      img_sonografia INTEGER DEFAULT 0,
      img_tac INTEGER DEFAULT 0,
      img_otras TEXT,
      img_indicar_parte_cuerpo TEXT,

      img_via_ind_ambulatoria INTEGER DEFAULT 0,
      img_via_ind_emergencias INTEGER DEFAULT 0,

      dx_1 TEXT,
      dx_2 TEXT,
      dx_3 TEXT,

      manejo_1 TEXT,
      manejo_2 TEXT,
      manejo_3 TEXT,

      observaciones TEXT,

      proc_sutura INTEGER DEFAULT 0,
      proc_inmovilizacion INTEGER DEFAULT 0,
      proc_reanimacion INTEGER DEFAULT 0,
      proc_nebulizacion INTEGER DEFAULT 0,
      proc_otros TEXT,

      interconsulta_no INTEGER DEFAULT 0,
      interconsulta_si INTEGER DEFAULT 0,
      interconsulta_especialidad TEXT,
      interconsulta_especialista TEXT,

      destino_admitido INTEGER DEFAULT 0,
      destino_fallecido INTEGER DEFAULT 0,
      destino_fuga INTEGER DEFAULT 0,
      destino_alta INTEGER DEFAULT 0,
      destino_alta_peticion INTEGER DEFAULT 0,
      destino_referido_a TEXT,

      observaciones_destino TEXT,

      medico_tratante TEXT,
      medico_exequatur TEXT
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_hc_aseg_estado ON hc_asegurados(estado)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_hc_aseg_created ON hc_asegurados(created_at)")

    def set_default(key, value):
        cur.execute("INSERT OR IGNORE INTO settings(key, value) VALUES (?,?)", (key, str(value)))

    set_default("branding.org_name", "Mirsa Tech Hospital")
    set_default("branding.powered_by", "Mirsa Tech®")
    set_default("branding.logo_path", "")
    set_default("enfermeria.limit", "20")

    # seed materials
    cur.execute("SELECT COUNT(*) AS c FROM materiales_catalogo")
    if int(cur.fetchone()["c"]) == 0:
        for m in [
            "Sol. Salina 1000ml",
            "Gasas estériles",
            "Vendajes",
            "Catéteres",
            "Suero fisiológico",
            "Jeringas",
            "Agujas",
        ]:
            cur.execute(
                "INSERT OR IGNORE INTO materiales_catalogo(nombre, activo, created_at) VALUES (?,?,?)",
                (m, 1, now_local()),
            )

    # seed users
    cur.execute("SELECT COUNT(*) AS c FROM users")
    if int(cur.fetchone()["c"]) == 0:
        seed_users = [
            ("emergencias", "Médico Emergencias", "1234", "emergencias"),
            ("especialista", "Médico Especialista", "1234", "especialista"),
            ("enfermeria", "Enfermería", "1234", "enfermeria"),
            ("farmacia", "Farmacia", "1234", "farmacia"),
            ("facturacion", "Facturación", "1234", "facturacion"),
            ("admision", "Admisión", "1234", "admision"),
            ("auditor", "Auditor", "1234", "auditor"),
            ("admin", "Administrador", "1234", "administrador"),
            ("propietario", "Propietario", "1234", "superadmin"),
            ("planta", "Médico de Planta", "1234", "planta"),
        ]
        for username, name, pwd, role in seed_users:
            cur.execute(
                """INSERT INTO users(username, name, password_hash, role, active, must_change_password, created_at)
                   VALUES (?,?,?,?,?,?,?)""",
                (username, name, generate_password_hash(pwd), role, 1, 0, now_local()),
            )

    conn.commit()
    conn.close()

# =========================
# Auth + Guards
# =========================

def require_login(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return jsonify({"ok": False, "error": "No autorizado"}), 401
        return fn(*args, **kwargs)
    return wrapper

def require_role(*roles):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            u = session.get("user")
            if not u:
                return jsonify({"ok": False, "error": "No autorizado"}), 401
            if u.get("role") not in roles:
                return jsonify({"ok": False, "error": "No permitido"}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator

# =========================
# Pages
# =========================

@APP.get("/")
def home():
    return redirect(url_for("index_page"))

@APP.get("/app")
def index_page():
    return render_template("index.html")

@APP.get("/branding/logo")
def branding_logo():
    logo_path = get_setting("branding.logo_path", "")
    if not logo_path:
        return ("", 404)
    full = os.path.join(BRANDING_DIR, logo_path)
    if not os.path.exists(full):
        return ("", 404)
    return send_file(full)

# =========================
# API: Branding
# =========================

@APP.get("/api/branding")
def api_branding():
    return jsonify({
        "ok": True,
        "org_name": get_setting("branding.org_name", "Mirsa Tech Hospital"),
        "powered_by": get_setting("branding.powered_by", "Mirsa Tech®"),
        "logo_url": "/branding/logo" if get_setting("branding.logo_path", "") else ""
    })

# =========================
# API: Auth
# =========================

@APP.post("/api/auth/login")
def api_login():
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    role = (data.get("role") or "").strip()

    if not username or not password or not role:
        return jsonify({"ok": False, "error": "Datos incompletos"}), 400

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, username, name, password_hash, role, active, must_change_password FROM users WHERE username=?", (username,))
    row = cur.fetchone()
    conn.close()

    if not row or int(row["active"]) != 1:
        return jsonify({"ok": False, "error": "Usuario inválido"}), 401
    if row["role"] != role:
        return jsonify({"ok": False, "error": "Rol incorrecto"}), 401
    if not check_password_hash(row["password_hash"], password):
        return jsonify({"ok": False, "error": "Contraseña incorrecta"}), 401

    session["user"] = {
        "id": row["id"],
        "username": row["username"],
        "name": row["name"] or row["username"],
        "role": row["role"],
        "must_change_password": int(row["must_change_password"] or 0),
    }
    audit("LOGIN", entity="user", entity_id=row["id"], after={"username": row["username"], "role": row["role"]})
    return jsonify({"ok": True, "user": session["user"]})

@APP.post("/api/auth/logout")
@require_login
def api_logout():
    u = session.get("user")
    audit("LOGOUT", entity="user", entity_id=u.get("id") if u else None, after={"username": u.get("username")} if u else None)
    session.pop("user", None)
    return jsonify({"ok": True})

@APP.get("/api/auth/me")
def api_me():
    return jsonify({"ok": True, "user": session.get("user")})

@APP.post("/api/auth/change_password")
@require_login
def api_change_password():
    data = request.get_json(force=True, silent=True) or {}
    old = (data.get("old_password") or "").strip()
    new = (data.get("new_password") or "").strip()
    if len(new) < 4:
        return jsonify({"ok": False, "error": "La nueva contraseña debe tener al menos 4 caracteres"}), 400

    u = session["user"]
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT password_hash, must_change_password FROM users WHERE id=?", (u["id"],))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Usuario no existe"}), 404

    must = int(row["must_change_password"] or 0)
    if must == 0:
        if not old:
            conn.close()
            return jsonify({"ok": False, "error": "Debes escribir tu contraseña actual"}), 400
        if not check_password_hash(row["password_hash"], old):
            conn.close()
            return jsonify({"ok": False, "error": "Contraseña actual incorrecta"}), 401

    cur.execute("UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?", (generate_password_hash(new), u["id"]))
    conn.commit()
    conn.close()

    audit("CHANGE_PASSWORD", entity="user", entity_id=u["id"], after={"username": u["username"]})
    u["must_change_password"] = 0
    session["user"] = u
    return jsonify({"ok": True})

# =========================
# Superadmin: DB Info + Reset DB
# =========================

@APP.get("/api/superadmin/db_info")
@require_role("superadmin")
def api_superadmin_db_info():
    return jsonify({
        "ok": True,
        "db_path": os.path.abspath(DB_PATH),
        "data_dir": os.path.abspath(DB_DIR),
        "exports_dir": os.path.abspath(EXPORTS_DIR),
        "backups_dir": os.path.abspath(BACKUPS_DIR),
    })

@APP.post("/api/superadmin/reset_db")
@require_role("superadmin")
def api_superadmin_reset_db():
    data = request.get_json(force=True, silent=True) or {}
    confirm = (data.get("confirm") or "").strip()
    if confirm != "RESET":
        return jsonify({"ok": False, "error": 'Confirmación inválida. Escribe exactamente: RESET'}), 400

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUPS_DIR, f"hospital_backup_{ts}.db")

    if os.path.exists(DB_PATH):
        shutil.copy2(DB_PATH, backup_path)

    for p in (DB_PATH, DB_PATH + "-wal", DB_PATH + "-shm"):
        try:
            if os.path.exists(p):
                os.remove(p)
        except Exception:
            pass

    init_db()
    audit("RESET_DB", entity="db", after={"backup_path": backup_path})
    return jsonify({"ok": True, "backup_path": backup_path})

# =========================
# API: Settings + Branding upload (Superadmin)
# =========================

@APP.put("/api/settings")
@require_role("superadmin")
def api_settings_put():
    data = request.get_json(force=True, silent=True) or {}
    updates = data.get("updates")
    if not isinstance(updates, dict):
        return jsonify({"ok": False, "error": "updates inválido"}), 400

    allowed_prefixes = ("enfermeria.", "branding.", "server.", "data.")
    before = {}
    after = {}

    for k, v in updates.items():
        if not isinstance(k, str):
            continue
        if not k.startswith(allowed_prefixes):
            continue
        before[k] = get_setting(k, "")
        set_setting(k, str(v))
        after[k] = str(v)

    audit("UPDATE_SETTINGS", entity="settings", before=before, after=after)
    return jsonify({"ok": True, "updated": after, "note": "Cambios de puerto/ruta requieren reinicio."})

@APP.post("/api/branding/logo")
@require_role("superadmin")
def api_branding_logo_upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Falta archivo"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "Nombre inválido"}), 400
    filename = secure_filename(f.filename)
    _, ext = os.path.splitext(filename)
    if ext.lower() not in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        return jsonify({"ok": False, "error": "Formato no soportado"}), 400

    final_name = f"logo{ext.lower()}"
    full = os.path.join(BRANDING_DIR, final_name)
    f.save(full)

    before = {"branding.logo_path": get_setting("branding.logo_path", "")}
    set_setting("branding.logo_path", final_name)
    after = {"branding.logo_path": final_name}
    audit("UPLOAD_LOGO", entity="branding", before=before, after=after)
    return jsonify({"ok": True, "logo_url": "/branding/logo"})

# =========================
# API: Materiales
# =========================

@APP.get("/api/materiales")
@require_login
def api_materiales_list():
    q = (request.args.get("q") or "").strip().lower()
    conn = db()
    cur = conn.cursor()
    if q:
        cur.execute("SELECT id, nombre, activo FROM materiales_catalogo WHERE lower(nombre) LIKE ? ORDER BY nombre", (f"%{q}%",))
    else:
        cur.execute("SELECT id, nombre, activo FROM materiales_catalogo ORDER BY nombre")
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})

@APP.post("/api/superadmin/materiales/add")
@require_role("superadmin")
def api_super_material_add():
    data = request.get_json(force=True, silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"ok": False, "error": "Nombre requerido"}), 400
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("INSERT INTO materiales_catalogo(nombre, activo, created_at) VALUES (?,?,?)", (nombre, 1, now_local()))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "Ese material ya existe"}), 400
    conn.close()
    audit("ADD_MATERIAL", entity="materiales_catalogo", after={"nombre": nombre})
    return jsonify({"ok": True})

@APP.post("/api/superadmin/materiales/<int:mid>/toggle_active")
@require_role("superadmin")
def api_super_material_toggle(mid: int):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT activo FROM materiales_catalogo WHERE id=?", (mid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Material no existe"}), 404
    new_active = 0 if int(row["activo"]) == 1 else 1
    cur.execute("UPDATE materiales_catalogo SET activo=? WHERE id=?", (new_active, mid))
    conn.commit()
    conn.close()
    audit("TOGGLE_MATERIAL_ACTIVE", entity="materiales_catalogo", entity_id=mid, before={"activo": int(row["activo"])}, after={"activo": new_active})
    return jsonify({"ok": True, "activo": new_active})

@APP.delete("/api/superadmin/materiales/<int:mid>")
@require_role("superadmin")
def api_super_material_delete(mid: int):
    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT nombre FROM materiales_catalogo WHERE id=?", (mid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Material no existe"}), 404

    try:
        cur.execute("DELETE FROM materiales_catalogo WHERE id=?", (mid,))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "No se puede eliminar: material en uso. (Usa desactivar si quieres ocultarlo.)"}), 400

    conn.close()
    audit("DELETE_MATERIAL", entity="materiales_catalogo", entity_id=mid, after={"nombre": row["nombre"]})
    return jsonify({"ok": True})

def _xlsx_required_file():
    if "file" not in request.files:
        return None, (jsonify({"ok": False, "error": "Falta archivo"}), 400)
    f = request.files["file"]
    if not f.filename:
        return None, (jsonify({"ok": False, "error": "Nombre inválido"}), 400)
    return f, None

@APP.get("/api/superadmin/templates/materiales.xlsx")
@require_role("superadmin")
def api_template_materials_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "materials"
    ws.append(["nombre", "activo"])
    ws.append(["Gasas estériles", 1])
    ws.append(["Jeringas 5ml", 1])
    out_path = os.path.join(EXPORTS_DIR, "materiales_template.xlsx")
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name="materiales_template.xlsx")

@APP.post("/api/superadmin/import/materiales")
@require_role("superadmin")
def api_import_materiales():
    f, err = _xlsx_required_file()
    if err:
        return err
    filename = secure_filename(f.filename)
    tmp = os.path.join(EXPORTS_DIR, f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
    f.save(tmp)

    wb = load_workbook(tmp)
    if "materials" not in wb.sheetnames:
        return jsonify({"ok": False, "error": "El Excel debe tener una hoja llamada 'materials'"}), 400

    ws = wb["materials"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    idx = {name: header.index(name) for name in header if name}

    if "nombre" not in idx:
        return jsonify({"ok": False, "error": "Falta columna requerida: nombre"}), 400

    created = 0
    updated = 0
    errors = []

    conn = db()
    cur = conn.cursor()

    for i, r in enumerate(rows[1:], start=2):
        try:
            nombre = (r[idx["nombre"]] or "").strip() if r[idx["nombre"]] is not None else ""
            if not nombre:
                raise ValueError("nombre vacío")
            activo_val = r[idx["activo"]] if "activo" in idx else 1
            activo = 1
            if activo_val is not None and str(activo_val).strip() != "":
                activo = 1 if str(activo_val).strip() in ("1", "true", "True", "SI", "Si", "si") else 0

            cur.execute("SELECT id FROM materiales_catalogo WHERE nombre=?", (nombre,))
            ex = cur.fetchone()
            if ex:
                cur.execute("UPDATE materiales_catalogo SET activo=? WHERE id=?", (activo, ex["id"]))
                updated += 1
            else:
                cur.execute("INSERT INTO materiales_catalogo(nombre, activo, created_at) VALUES (?,?,?)", (nombre, activo, now_local()))
                created += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    conn.commit()
    conn.close()

    audit("IMPORT_MATERIALES", entity="materiales_catalogo", after={"created": created, "updated": updated, "errors": len(errors)})
    return jsonify({"ok": True, "created": created, "updated": updated, "errors": errors})

# =========================
# API: Users (Superadmin)
# =========================

@APP.get("/api/superadmin/users")
@require_role("superadmin")
def api_super_users_list():
    q = (request.args.get("q") or "").strip().lower()
    conn = db()
    cur = conn.cursor()
    if q:
        cur.execute("""
          SELECT id, username, name, role, active, must_change_password, created_at, exequatur
          FROM users
          WHERE lower(username) LIKE ? OR lower(name) LIKE ?
          ORDER BY id DESC
          LIMIT 500
        """, (f"%{q}%", f"%{q}%"))
    else:
        cur.execute("""
          SELECT id, username, name, role, active, must_change_password, created_at, exequatur
          FROM users
          ORDER BY id DESC
          LIMIT 500
        """)
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})

@APP.post("/api/superadmin/users")
@require_role("superadmin")
def api_super_users_create():
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip()
    name = (data.get("name") or "").strip()
    role = (data.get("role") or "").strip().lower()
    password = (data.get("password") or "").strip()
    exequatur = (data.get("exequatur") or "").strip()

    if not username or not role:
        return jsonify({"ok": False, "error": "username y role son requeridos"}), 400
    if role == "superadmin":
        return jsonify({"ok": False, "error": "No se permite crear/importar superadmin"}), 400
    if role not in ROLES:
        return jsonify({"ok": False, "error": "role inválido"}), 400

    if not password:
        password = "1234"
    must_change = 1

    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("""
          INSERT INTO users(username, name, password_hash, role, active, must_change_password, created_at, exequatur)
          VALUES (?,?,?,?,1,?,?,?,?)
        """, (username, name, generate_password_hash(password), role, must_change, now_local(), exequatur))
        uid = cur.lastrowid
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "Ese username ya existe"}), 400

    conn.close()
    audit("CREATE_USER", entity="user", entity_id=uid, after={"username": username, "role": role, "exequatur": exequatur})
    return jsonify({"ok": True, "id": uid})

@APP.post("/api/superadmin/users/<int:uid>/set_exequatur")
@require_role("superadmin")
def api_superadmin_set_exequatur(uid: int):
    data = request.get_json(force=True, silent=True) or {}
    ex = (data.get("exequatur") or "").strip()

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT username FROM users WHERE id=?", (uid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Usuario no existe"}), 404

    cur.execute("UPDATE users SET exequatur=? WHERE id=?", (ex, uid))
    conn.commit()
    conn.close()

    audit("SET_EXEQUATUR", entity="user", entity_id=uid, after={"username": row["username"], "exequatur": ex})
    return jsonify({"ok": True})

@APP.post("/api/superadmin/users/<int:uid>/toggle_active")
@require_role("superadmin")
def api_super_users_toggle(uid: int):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT active FROM users WHERE id=?", (uid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Usuario no existe"}), 404
    new_active = 0 if int(row["active"]) == 1 else 1
    cur.execute("UPDATE users SET active=? WHERE id=?", (new_active, uid))
    conn.commit()
    conn.close()
    audit("TOGGLE_USER_ACTIVE", entity="user", entity_id=uid, before={"active": int(row["active"])}, after={"active": new_active})
    return jsonify({"ok": True, "active": new_active})

@APP.post("/api/superadmin/users/<int:uid>/reset_password_1234")
@require_role("superadmin")
def api_super_users_reset(uid: int):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT username, role FROM users WHERE id=?", (uid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Usuario no existe"}), 404
    if row["role"] == "superadmin":
        conn.close()
        return jsonify({"ok": False, "error": "No se permite resetear superadmin aquí"}), 400

    cur.execute("UPDATE users SET password_hash=?, must_change_password=1 WHERE id=?", (generate_password_hash("1234"), uid))
    conn.commit()
    conn.close()
    audit("RESET_PASSWORD", entity="user", entity_id=uid, after={"username": row["username"], "password": "1234"})
    return jsonify({"ok": True})

@APP.delete("/api/superadmin/users/<int:uid>")
@require_role("superadmin")
def api_super_users_delete(uid: int):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT username, role FROM users WHERE id=?", (uid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Usuario no existe"}), 404
    if row["role"] == "superadmin":
        conn.close()
        return jsonify({"ok": False, "error": "No se permite eliminar usuarios superadmin"}), 400

    username = row["username"]
    cur.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()

    audit("DELETE_USER", entity="user", entity_id=uid, after={"username": username})
    return jsonify({"ok": True})

@APP.get("/api/superadmin/templates/users.xlsx")
@require_role("superadmin")
def api_template_users_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["username", "name", "role", "password", "active", "exequatur"])
    ws.append(["emergencias3", "Dr. Juan Pérez", "emergencias", "1234", 1, "12345"])
    ws.append(["enfermeria2", "Ana Gómez", "enfermeria", "", 1, ""])
    out_path = os.path.join(EXPORTS_DIR, "usuarios_template.xlsx")
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name="usuarios_template.xlsx")

@APP.post("/api/superadmin/import/users")
@require_role("superadmin")
def api_import_users():
    f, err = _xlsx_required_file()
    if err:
        return err
    filename = secure_filename(f.filename)
    tmp = os.path.join(EXPORTS_DIR, f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
    f.save(tmp)

    wb = load_workbook(tmp)
    if "users" not in wb.sheetnames:
        return jsonify({"ok": False, "error": "El Excel debe tener una hoja llamada 'users'"}), 400

    ws = wb["users"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    idx = {name: header.index(name) for name in header if name}

    for required in ("username", "role"):
        if required not in idx:
            return jsonify({"ok": False, "error": f"Falta columna requerida: {required}"}), 400

    created = 0
    updated = 0
    errors = []

    conn = db()
    cur = conn.cursor()

    for i, r in enumerate(rows[1:], start=2):
        try:
            username = (r[idx["username"]] or "").strip() if r[idx["username"]] is not None else ""
            role = (r[idx["role"]] or "").strip().lower() if r[idx["role"]] is not None else ""
            name = (r[idx["name"]] or "").strip() if "name" in idx and r[idx["name"]] is not None else ""
            password = (r[idx["password"]] or "").strip() if "password" in idx and r[idx["password"]] is not None else ""
            active_val = r[idx["active"]] if "active" in idx else 1
            exequatur = (r[idx["exequatur"]] or "").strip() if "exequatur" in idx and r[idx["exequatur"]] is not None else ""

            if not username:
                raise ValueError("username vacío")
            if role == "superadmin":
                raise ValueError("No se permite importar role=superadmin")
            if role not in ROLES:
                raise ValueError(f"role inválido: {role}")

            active = 1
            if active_val is not None and str(active_val).strip() != "":
                active = 1 if str(active_val).strip() in ("1", "true", "True", "SI", "Si", "si") else 0

            if not password:
                password = "1234"
            must_change = 1

            cur.execute("SELECT id FROM users WHERE username=?", (username,))
            ex = cur.fetchone()
            if ex:
                cur.execute("""
                  UPDATE users
                  SET name=?, role=?, active=?,
                      password_hash=?,
                      must_change_password=?,
                      exequatur=?
                  WHERE username=?
                """, (name, role, active, generate_password_hash(password), must_change, exequatur, username))
                updated += 1
            else:
                cur.execute("""
                  INSERT INTO users(username, name, password_hash, role, active, must_change_password, created_at, exequatur)
                  VALUES (?,?,?,?,?,?,?,?)
                """, (username, name, generate_password_hash(password), role, active, must_change, now_local(), exequatur))
                created += 1

        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    conn.commit()
    conn.close()

    audit("IMPORT_USERS", entity="users", after={"created": created, "updated": updated, "errors": len(errors)})
    return jsonify({"ok": True, "created": created, "updated": updated, "errors": errors})

# =========================
# API: Emergencias (normal)
# =========================

@APP.post("/api/emergencias/pacientes")
@require_role("emergencias")
def api_emergencias_crear():
    data = request.get_json(force=True, silent=True) or {}
    paciente = (data.get("paciente") or "").strip()
    diagnostico = (data.get("diagnostico") or "").strip()
    tratamiento = (data.get("tratamiento") or "").strip()
    if not paciente or not diagnostico or not tratamiento:
        return jsonify({"ok": False, "error": "Campos incompletos"}), 400

    conn = db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO pacientes(created_at, paciente, diagnostico, tratamiento, tipo, estado, medico_user)
           VALUES (?,?,?,?, 'emergencias', 'pendiente', ?)""",
        (now_local(), paciente, diagnostico, tratamiento, session["user"]["username"]),
    )
    pid = cur.lastrowid
    conn.commit()
    conn.close()
    audit("CREATE_PATIENT", entity="pacientes", entity_id=pid, after={"paciente": paciente, "tipo": "emergencias"})
    return jsonify({"ok": True, "id": pid})

@APP.get("/api/emergencias/pacientes")
@require_role("emergencias")
def api_emergencias_listar():
    try:
        limit = int(request.args.get("limit") or "10")
    except Exception:
        limit = 10
    limit = max(1, min(limit, 200))
    me = session["user"]["username"]

    conn = db()
    cur = conn.cursor()
    cur.execute(
        """SELECT id, created_at, paciente, diagnostico, tratamiento, estado, enfermera_user, aplicado_at
           FROM pacientes
           WHERE medico_user=?
           ORDER BY id DESC
           LIMIT ?""",
        (me, limit),
    )
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})

@APP.put("/api/emergencias/pacientes/<int:pid>")
@require_role("emergencias")
def api_emergencias_editar(pid: int):
    data = request.get_json(force=True, silent=True) or {}
    paciente = (data.get("paciente") or "").strip()
    diagnostico = (data.get("diagnostico") or "").strip()
    tratamiento = (data.get("tratamiento") or "").strip()
    if not paciente or not diagnostico or not tratamiento:
        return jsonify({"ok": False, "error": "Campos incompletos"}), 400

    me = session["user"]["username"]
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT paciente, diagnostico, tratamiento, estado FROM pacientes WHERE id=? AND medico_user=?", (pid, me))
    prev = cur.fetchone()
    if not prev:
        conn.close()
        return jsonify({"ok": False, "error": "No existe o no pertenece a tu usuario"}), 404

    before = dict(prev)
    cur.execute(
        "UPDATE pacientes SET paciente=?, diagnostico=?, tratamiento=? WHERE id=? AND medico_user=?",
        (paciente, diagnostico, tratamiento, pid, me),
    )
    conn.commit()
    conn.close()
    audit("UPDATE_PATIENT", entity="pacientes", entity_id=pid, before=before, after={"paciente": paciente, "diagnostico": diagnostico})
    return jsonify({"ok": True})

@APP.get("/api/emergencias/export.xlsx")
@require_role("emergencias")
def api_emergencias_export_excel():
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()
    me = session["user"]["username"]

    conn = db()
    cur = conn.cursor()
    sql = """SELECT created_at, paciente, diagnostico, tratamiento, estado, enfermera_user
             FROM pacientes
             WHERE medico_user=?"""
    params = [me]
    if desde:
        sql += " AND substr(created_at,1,10) >= ?"
        params.append(desde)
    if hasta:
        sql += " AND substr(created_at,1,10) <= ?"
        params.append(hasta)
    sql += " ORDER BY id DESC"
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pacientes"
    ws.append(["Fecha", "Paciente", "Diagnóstico", "Tratamiento", "Estado", "Enfermera"])
    for r in rows:
        ws.append([
            r["created_at"],
            r["paciente"],
            r["diagnostico"],
            r["tratamiento"],
            "Aplicado" if r["estado"] == "aplicado" else "Pendiente",
            r["enfermera_user"] or ""
        ])

    out_path = os.path.join(EXPORTS_DIR, f"mis_pacientes_{me}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(out_path)
    audit("EXPORT_EXCEL", entity="pacientes", after={"desde": desde, "hasta": hasta, "count": len(rows)})
    return send_file(out_path, as_attachment=True, download_name=f"mis_pacientes_{me}.xlsx")

# =========================
# API: Enfermería
# =========================

@APP.get("/api/enfermeria/pacientes")
@require_role("enfermeria")
def api_enfermeria_listar():
    try:
        limit = int(get_setting("enfermeria.limit", "20"))
    except Exception:
        limit = 20
    limit = max(5, min(limit, 200))

    conn = db()
    cur = conn.cursor()
    cur.execute(
        """SELECT id, created_at, paciente, diagnostico, tratamiento, tipo, estado, medico_user, enfermera_user, aplicado_at
           FROM pacientes
           ORDER BY id DESC
           LIMIT ?""",
        (limit,),
    )
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items, "limit": limit})

@APP.get("/api/enfermeria/pacientes/<int:pid>/materiales")
@require_role("enfermeria")
def api_enfermeria_get_materiales(pid: int):
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT mp.material_id, mc.nombre AS material_nombre, mp.cantidad
      FROM materiales_paciente mp
      JOIN materiales_catalogo mc ON mc.id = mp.material_id
      WHERE mp.paciente_id=?
      ORDER BY mc.nombre
    """, (pid,))
    items = [{"material_id": r["material_id"], "material": r["material_nombre"], "cantidad": r["cantidad"]} for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})

@APP.post("/api/enfermeria/pacientes/<int:pid>/materiales")
@require_role("enfermeria")
def api_enfermeria_set_materiales(pid: int):
    data = request.get_json(force=True, silent=True) or {}
    items = data.get("items")
    if not isinstance(items, list):
        return jsonify({"ok": False, "error": "items inválido"}), 400

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM pacientes WHERE id=?", (pid,))
    if not cur.fetchone():
        conn.close()
        return jsonify({"ok": False, "error": "Paciente no existe"}), 404

    cur.execute("SELECT material_id, cantidad FROM materiales_paciente WHERE paciente_id=?", (pid,))
    before = {str(r["material_id"]): int(r["cantidad"]) for r in cur.fetchall()}

    user = session["user"]["username"]
    for it in items:
        try:
            material_id = int(it.get("material_id"))
            cantidad = int(it.get("cantidad"))
        except Exception:
            continue
        if cantidad <= 0:
            cur.execute("DELETE FROM materiales_paciente WHERE paciente_id=? AND material_id=?", (pid, material_id))
            continue

        cur.execute("""
          INSERT INTO materiales_paciente(paciente_id, material_id, cantidad, updated_at, updated_by)
          VALUES (?,?,?,?,?)
          ON CONFLICT(paciente_id, material_id)
          DO UPDATE SET cantidad=excluded.cantidad, updated_at=excluded.updated_at, updated_by=excluded.updated_by
        """, (pid, material_id, cantidad, now_local(), user))

    conn.commit()

    cur.execute("SELECT material_id, cantidad FROM materiales_paciente WHERE paciente_id=?", (pid,))
    after = {str(r["material_id"]): int(r["cantidad"]) for r in cur.fetchall()}
    conn.close()

    audit("UPDATE_MATERIALS", entity="pacientes", entity_id=pid, before={"materiales": before}, after={"materiales": after})
    return jsonify({"ok": True})

@APP.post("/api/enfermeria/pacientes/<int:pid>/aplicar")
@require_role("enfermeria")
def api_enfermeria_aplicar(pid: int):
    user = session["user"]["username"]
    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT estado, enfermera_user FROM pacientes WHERE id=?", (pid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "No existe"}), 404

    cur.execute("SELECT COUNT(*) AS c FROM materiales_paciente WHERE paciente_id=?", (pid,))
    if int(cur.fetchone()["c"] or 0) <= 0:
        conn.close()
        return jsonify({"ok": False, "error": "Debes seleccionar al menos 1 material gastable antes de aplicar."}), 400

    before = dict(row)
    cur.execute("UPDATE pacientes SET estado='aplicado', enfermera_user=?, aplicado_at=? WHERE id=?", (user, now_local(), pid))
    conn.commit()
    conn.close()

    audit("APPLY_PATIENT", entity="pacientes", entity_id=pid, before=before, after={"estado": "aplicado", "enfermera_user": user})
    return jsonify({"ok": True})

@APP.post("/api/enfermeria/ambulatorio")
@require_role("enfermeria")
def api_enfermeria_ambulatorio():
    data = request.get_json(force=True, silent=True) or {}
    paciente = (data.get("paciente") or "").strip()
    medicamento = (data.get("medicamento") or "").strip()
    material_items = data.get("materiales")

    if not paciente or not medicamento:
        return jsonify({"ok": False, "error": "Paciente y medicamento son requeridos"}), 400
    if not isinstance(material_items, list) or len(material_items) == 0:
        return jsonify({"ok": False, "error": "Ambulatorio requiere al menos 1 material gastable."}), 400

    enf_user = session["user"]["username"]
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      INSERT INTO pacientes(created_at, paciente, diagnostico, tratamiento, tipo, estado, medico_user, enfermera_user, aplicado_at)
      VALUES (?,?,?,?, 'ambulatorio', 'aplicado', NULL, ?, ?)
    """, (now_local(), paciente, "Ambulatorio", medicamento, enf_user, now_local()))
    pid = cur.lastrowid

    for it in material_items:
        try:
            material_id = int(it.get("material_id"))
            cantidad = int(it.get("cantidad"))
        except Exception:
            continue
        if cantidad <= 0:
            continue
        cur.execute("""
          INSERT INTO materiales_paciente(paciente_id, material_id, cantidad, updated_at, updated_by)
          VALUES (?,?,?,?,?)
          ON CONFLICT(paciente_id, material_id)
          DO UPDATE SET cantidad=excluded.cantidad, updated_at=excluded.updated_at, updated_by=excluded.updated_by
        """, (pid, material_id, cantidad, now_local(), enf_user))

    conn.commit()
    conn.close()
    audit("CREATE_AMBULATORIO", entity="pacientes", entity_id=pid, after={"paciente": paciente, "medicamento": medicamento})
    return jsonify({"ok": True, "id": pid})

# =========================
# API: Audit log
# =========================

@APP.get("/api/audit")
@require_role("superadmin", "auditor")
def api_audit_list():
    try:
        limit = int(request.args.get("limit") or "200")
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT id, created_at, user, role, action, entity, entity_id, before_json, after_json
      FROM audit_log
      ORDER BY id DESC
      LIMIT ?
    """, (limit,))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})

# =========================
# NUEVO: pacientes master + asegurados endpoints
# =========================

FACT_FIELDS_UP_TO_MOTIVO = {
    "fecha", "hora_llegada",
    "nombre", "apodo", "sexo", "edad", "fecha_nacimiento", "cedula_pasaporte_nui",
    "aseguradora", "nss", "grupo_sanguineo", "alergico", "direccion", "nacionalidad", "telefono",
    "via_llegada", "ambulancia_no", "paramedico", "acompanante", "parentesco", "acompanante_telefono", "acompanante_direccion",
    "motivo_cefalea", "motivo_tos", "motivo_palpitaciones", "motivo_hematemesis", "motivo_dolor_lumbar", "motivo_diarrea",
    "motivo_mareos", "motivo_epistaxis", "motivo_nauseas", "motivo_dolor_toracico", "motivo_hematuria", "motivo_herida_arma",
    "motivo_fiebre", "motivo_disnea", "motivo_vomitos", "motivo_dolor_abdominal", "motivo_otro"
}

@APP.get("/api/pacientes_master")
@require_login
def api_pacientes_master_search():
    q = (request.args.get("q") or "").strip().lower()
    if not q or len(q) < 2:
        return jsonify({"ok": True, "items": []})

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT id, nombre, apodo, sexo, edad, fecha_nacimiento, cedula_pasaporte_nui, nss,
             aseguradora, grupo_sanguineo, alergico, direccion, nacionalidad, telefono
      FROM pacientes_master
      WHERE lower(nombre) LIKE ?
         OR lower(cedula_pasaporte_nui) LIKE ?
         OR lower(telefono) LIKE ?
      ORDER BY id DESC
      LIMIT 20
    """, (f"%{q}%", f"%{q}%", f"%{q}%"))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})

def upsert_paciente_master_from_hc(data: dict):
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return None

    ced = (data.get("cedula_pasaporte_nui") or "").strip()
    tel = (data.get("telefono") or "").strip()
    fn = (data.get("fecha_nacimiento") or "").strip()

    payload = {
        "nombre": nombre,
        "apodo": (data.get("apodo") or "").strip(),
        "sexo": (data.get("sexo") or "").strip(),
        "edad": (data.get("edad") or "").strip(),
        "fecha_nacimiento": fn,
        "cedula_pasaporte_nui": ced,
        "nss": (data.get("nss") or "").strip(),
        "aseguradora": (data.get("aseguradora") or "").strip(),
        "grupo_sanguineo": (data.get("grupo_sanguineo") or "").strip(),
        "alergico": (data.get("alergico") or "").strip(),
        "direccion": (data.get("direccion") or "").strip(),
        "nacionalidad": (data.get("nacionalidad") or "").strip(),
        "telefono": tel,
    }

    conn = db()
    cur = conn.cursor()

    existing = None
    if ced:
        cur.execute("SELECT id FROM pacientes_master WHERE cedula_pasaporte_nui=? ORDER BY id DESC LIMIT 1", (ced,))
        existing = cur.fetchone()

    if not existing and nombre and fn and tel:
        cur.execute("""
          SELECT id FROM pacientes_master
          WHERE nombre=? AND fecha_nacimiento=? AND telefono=?
          ORDER BY id DESC LIMIT 1
        """, (nombre, fn, tel))
        existing = cur.fetchone()

    if existing:
        pid = existing["id"]
        cur.execute("""
          UPDATE pacientes_master
          SET updated_at=?,
              nombre=?, apodo=?, sexo=?, edad=?, fecha_nacimiento=?, cedula_pasaporte_nui=?,
              nss=?, aseguradora=?, grupo_sanguineo=?, alergico=?, direccion=?, nacionalidad=?, telefono=?
          WHERE id=?
        """, (
            now_local(),
            payload["nombre"], payload["apodo"], payload["sexo"], payload["edad"], payload["fecha_nacimiento"], payload["cedula_pasaporte_nui"],
            payload["nss"], payload["aseguradora"], payload["grupo_sanguineo"], payload["alergico"], payload["direccion"], payload["nacionalidad"], payload["telefono"],
            pid
        ))
        conn.commit()
        conn.close()
        return pid

    cur.execute("""
      INSERT INTO pacientes_master(
        created_at, updated_at,
        nombre, apodo, sexo, edad, fecha_nacimiento, cedula_pasaporte_nui,
        nss, aseguradora, grupo_sanguineo, alergico, direccion, nacionalidad, telefono
      ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        now_local(), now_local(),
        payload["nombre"], payload["apodo"], payload["sexo"], payload["edad"], payload["fecha_nacimiento"], payload["cedula_pasaporte_nui"],
        payload["nss"], payload["aseguradora"], payload["grupo_sanguineo"], payload["alergico"], payload["direccion"], payload["nacionalidad"], payload["telefono"],
    ))
    pid = cur.lastrowid
    conn.commit()
    conn.close()
    return pid

def generate_expediente_for_now():
    dt = datetime.now()
    key = dt if dt.hour >= 8 else (dt - timedelta(days=1))
    day_key = key.strftime("%Y%m%d")

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT COUNT(*) AS c
      FROM hc_asegurados
      WHERE expediente_clinico LIKE ?
    """, (f"{day_key}-%",))
    c = int(cur.fetchone()["c"] or 0)
    conn.close()
    seq = c + 1
    return f"{day_key}-{seq:02d}"

@APP.post("/api/asegurados")
@require_role("facturacion")
def api_aseg_create():
    data = request.get_json(force=True, silent=True) or {}
    record = {k: data.get(k) for k in FACT_FIELDS_UP_TO_MOTIVO}

    expediente = generate_expediente_for_now()

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      INSERT INTO hc_asegurados(created_at, updated_at, estado, fact_user, expediente_clinico)
      VALUES (?,?, 'pendiente_medico', ?, ?)
    """, (now_local(), now_local(), session["user"]["username"], expediente))
    hid = cur.lastrowid

    sets = []
    vals = []
    for k, v in record.items():
        sets.append(f"{k}=?")
        vals.append(v)
    if sets:
        cur.execute(f"UPDATE hc_asegurados SET {', '.join(sets)}, updated_at=? WHERE id=?", (*vals, now_local(), hid))

    conn.commit()
    conn.close()

    upsert_paciente_master_from_hc(record)
    audit("ASEG_CREATE", entity="hc_asegurados", entity_id=hid, after={"fact_user": session["user"]["username"], "expediente": expediente})
    return jsonify({"ok": True, "id": hid, "expediente_clinico": expediente})

@APP.get("/api/asegurados")
@require_login
def api_aseg_list():
    role = session["user"]["role"]
    q = (request.args.get("q") or "").strip().lower()
    estado = (request.args.get("estado") or "").strip()
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()

    conn = db()
    cur = conn.cursor()

    sql = """SELECT id, created_at, updated_at, estado, nombre, fact_user, medico_user, fecha, expediente_clinico
             FROM hc_asegurados WHERE 1=1"""
    params = []

    if role == "facturacion":
        sql += " AND fact_user=?"
        params.append(session["user"]["username"])
    elif role == "emergencias":
        if not estado:
            estado = "pendiente_medico"
    elif role in ("superadmin", "auditor", "administrador"):
        pass
    else:
        conn.close()
        return jsonify({"ok": False, "error": "No permitido"}), 403

    if estado:
        sql += " AND estado=?"
        params.append(estado)

    if desde:
        sql += " AND substr(created_at,1,10) >= ?"
        params.append(desde)
    if hasta:
        sql += " AND substr(created_at,1,10) <= ?"
        params.append(hasta)

    if q:
        sql += " AND (lower(nombre) LIKE ? OR lower(fact_user) LIKE ? OR lower(expediente_clinico) LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

    sql += " ORDER BY id DESC LIMIT 300"
    cur.execute(sql, params)
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})

@APP.get("/api/asegurados/<int:hid>")
@require_login
def api_aseg_get(hid: int):
    role = session["user"]["role"]
    user = session["user"]["username"]

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM hc_asegurados WHERE id=?", (hid,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": False, "error": "No existe"}), 404

    if role == "facturacion" and row["fact_user"] != user:
        return jsonify({"ok": False, "error": "No permitido"}), 403
    if role not in ("facturacion", "emergencias", "superadmin", "auditor", "administrador"):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    doctor = None
    if row["medico_user"]:
        conn = db()
        cur = conn.cursor()
        cur.execute("SELECT username, name, exequatur FROM users WHERE username=?", (row["medico_user"],))
        d = cur.fetchone()
        conn.close()
        if d:
            doctor = {"username": d["username"], "name": d["name"], "exequatur": d["exequatur"]}

    return jsonify({"ok": True, "item": dict(row), "doctor": doctor})

@APP.put("/api/asegurados/<int:hid>")
@require_login
def api_aseg_update(hid: int):
    data = request.get_json(force=True, silent=True) or {}
    role = session["user"]["role"]
    user = session["user"]["username"]

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, fact_user FROM hc_asegurados WHERE id=?", (hid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "No existe"}), 404

    if role == "facturacion":
        if row["fact_user"] != user:
            conn.close()
            return jsonify({"ok": False, "error": "No permitido"}), 403
        allowed = FACT_FIELDS_UP_TO_MOTIVO
        data.pop("triaje_prioridad", None)  # triage no facturación
    elif role in ("emergencias", "auditor", "superadmin", "administrador"):
        allowed = None
    else:
        conn.close()
        return jsonify({"ok": False, "error": "No permitido"}), 403

    sets = []
    vals = []
    for k, v in data.items():
        if not isinstance(k, str):
            continue
        if allowed is not None and k not in allowed:
            continue
        sets.append(f"{k}=?")
        vals.append(v)

    if role == "emergencias":
        sets.append("medico_user=?")
        vals.append(user)

    if sets:
        cur.execute(f"UPDATE hc_asegurados SET {', '.join(sets)}, updated_at=? WHERE id=?", (*vals, now_local(), hid))
        conn.commit()

    conn.close()
    upsert_paciente_master_from_hc(data)
    audit("ASEG_UPDATE", entity="hc_asegurados", entity_id=hid, after={"role": role, "user": user})
    return jsonify({"ok": True})

@APP.post("/api/asegurados/<int:hid>/cerrar")
@require_role("emergencias")
def api_aseg_close(hid: int):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM hc_asegurados WHERE id=?", (hid,))
    if not cur.fetchone():
        conn.close()
        return jsonify({"ok": False, "error": "No existe"}), 404

    conn2 = db()
    cur2 = conn2.cursor()
    cur2.execute("SELECT name, exequatur FROM users WHERE username=?", (session["user"]["username"],))
    d = cur2.fetchone()
    conn2.close()

    medico_name = d["name"] if d else session["user"]["username"]
    medico_ex = d["exequatur"] if d else ""

    cur.execute("""
      UPDATE hc_asegurados
      SET estado='completo',
          medico_user=?,
          medico_tratante=?,
          medico_exequatur=?,
          updated_at=?
      WHERE id=?
    """, (session["user"]["username"], medico_name, medico_ex, now_local(), hid))
    conn.commit()
    conn.close()

    audit("ASEG_CLOSE", entity="hc_asegurados", entity_id=hid, after={"medico_user": session["user"]["username"]})
    return jsonify({"ok": True})
    
@APP.get("/api/users")
@require_login
def api_users_list_basic():
    role = (request.args.get("role") or "").strip().lower()
    if role and role not in ROLES:
        return jsonify({"ok": False, "error": "role inválido"}), 400

    conn = db()
    cur = conn.cursor()
    if role:
        cur.execute("""
          SELECT id, username, name, role, active, exequatur
          FROM users
          WHERE role=? AND active=1
          ORDER BY name, username
        """, (role,))
    else:
        cur.execute("""
          SELECT id, username, name, role, active, exequatur
          FROM users
          WHERE active=1
          ORDER BY role, name, username
        """)
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})
def _can_ingresos_read(role: str) -> bool:
    # lectura (todos los involucrados + auditoría)
    return role in ("emergencias", "planta", "especialista", "admision", "auditor", "administrador", "superadmin")

def _can_ingresos_write_medico(role: str) -> bool:
    # escritura médica (ingreso clínico, evoluciones, ordenes, epicrisis)
    return role in ("emergencias", "planta", "especialista", "auditor", "administrador", "superadmin")

def _can_historia_edit(iid: int, u: dict) -> bool:
    if u.get("role") == "auditor":
        return True
    # Solo médicos pueden ser "creador"
    if not _can_ingresos_write_medico(u.get("role", "")):
        return False

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT creado_por_user FROM ingresos WHERE id=?", (iid,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return False
    return (row["creado_por_user"] == u.get("username"))

def _can_ingresos_write_admision(role: str) -> bool:
    # hoja admisión
    return role in ("admision", "auditor", "administrador", "superadmin")

@APP.get("/api/ingresos/<int:iid>/historia")
@require_login
def api_ingresos_get_historia(iid: int):
    u = session["user"]
    # lectura: médicos y auditoría (ajusta si quieres permitir más)
    if u["role"] not in ["auditor","emergencias","planta","especialista","administrador","superadmin"]:
        return jsonify({"ok": False, "error": "No permitido"}), 403

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM ingresos_historia_clinica WHERE ingreso_id=?", (iid,))
    row = cur.fetchone()
    conn.close()
    return jsonify({"ok": True, "historia": dict(row) if row else None})

@APP.post("/api/ingresos")
@require_login
def api_ingresos_create():
    u = session["user"]
    if u["role"] not in ("emergencias", "planta", "especialista"):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    data = request.get_json(force=True, silent=True) or {}

    sala = (data.get("sala") or "").strip()
    habitacion = (data.get("habitacion") or "").strip()
    cama = (data.get("cama") or "").strip()
    especialidad = (data.get("especialidad") or "").strip()
    fecha_ingreso = (data.get("fecha_ingreso") or "").strip()
    hora_ingreso = (data.get("hora_ingreso") or "").strip()
    especialista_user = (data.get("especialista_user") or "").strip()
    if not especialista_user:
        return jsonify({"ok": False, "error": "Debes seleccionar el especialista que ordena el ingreso."}), 400

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      INSERT INTO ingresos(
        created_at, updated_at, estado,
        creado_por_user, creado_por_role,
        especialista_user,
        sala, habitacion, cama,
        especialidad, fecha_ingreso, hora_ingreso
      ) VALUES (?,?, 'activo', ?,?,?, ?,?,?, ?,?,?)
    """, (
        now_local(), now_local(),
        u["username"], u["role"],
        especialista_user,
        sala, habitacion, cama,
        especialidad, fecha_ingreso, hora_ingreso
    ))
    iid = cur.lastrowid

    # crear fila vacía de admisión para que admisión la complete
    cur.execute("""
      INSERT OR IGNORE INTO ingresos_admision(ingreso_id, updated_at, updated_by_user, updated_by_role)
      VALUES (?,?,?,?)
    """, (iid, now_local(), u["username"], u["role"]))

    conn.commit()
    conn.close()

    audit("CREATE_INGRESO", entity="ingresos", entity_id=iid, after={
        "creado_por_user": u["username"],
        "creado_por_role": u["role"],
        "especialista_user": especialista_user,
        "sala": sala, "habitacion": habitacion, "cama": cama,
        "especialidad": especialidad,
        "fecha_ingreso": fecha_ingreso,
        "hora_ingreso": hora_ingreso,
    })
    return jsonify({"ok": True, "id": iid})


@APP.get("/api/ingresos")
@require_login
def api_ingresos_list():
    u = session["user"]
    if not _can_ingresos_read(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    estado = (request.args.get("estado") or "activo").strip().lower()
    if estado not in ("activo", "egresado", "all"):
        return jsonify({"ok": False, "error": "estado inválido"}), 400

    conn = db()
    cur = conn.cursor()
    sql = """
      SELECT id, created_at, updated_at, estado,
             creado_por_user, creado_por_role, especialista_user,
             sala, habitacion, cama,
             especialidad, fecha_ingreso, hora_ingreso
      FROM ingresos
      WHERE 1=1
    """
    params = []
    if estado != "all":
        sql += " AND estado=?"
        params.append(estado)

    sql += " ORDER BY id DESC LIMIT 300"
    cur.execute(sql, params)
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "items": items})


@APP.get("/api/ingresos/<int:iid>")
@require_login
def api_ingresos_get(iid: int):
    u = session["user"]
    if not _can_ingresos_read(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM ingresos WHERE id=?", (iid,))
    ingreso = cur.fetchone()
    if not ingreso:
        conn.close()
        return jsonify({"ok": False, "error": "No existe"}), 404

    cur.execute("SELECT * FROM ingresos_admision WHERE ingreso_id=?", (iid,))
    adm = cur.fetchone()

    cur.execute("""
      SELECT * FROM ingreso_evoluciones
      WHERE ingreso_id=?
      ORDER BY datetime(fecha_hora) DESC, id DESC
      LIMIT 500
    """, (iid,))
    evoluciones = [dict(r) for r in cur.fetchall()]

    cur.execute("""
      SELECT * FROM ingreso_ordenes
      WHERE ingreso_id=?
      ORDER BY datetime(fecha_hora) DESC, id DESC
      LIMIT 500
    """, (iid,))
    ordenes = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT * FROM ingreso_epicrisis WHERE ingreso_id=?", (iid,))
    epi = cur.fetchone()

    conn.close()
    return jsonify({
        "ok": True,
        "ingreso": dict(ingreso),
        "admision": dict(adm) if adm else None,
        "evoluciones": evoluciones,
        "ordenes": ordenes,
        "epicrisis": dict(epi) if epi else None
    })
@APP.put("/api/ingresos/<int:iid>/historia")
@require_login
def api_ingresos_put_historia(iid: int):
    u = session["user"]
    if not _can_historia_edit(iid, u):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    data = request.get_json(force=True, silent=True) or {}

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT id FROM ingresos_historia_clinica WHERE ingreso_id=?", (iid,))
    prev = cur.fetchone()

    now = now_local()

    # columnas permitidas (white-list para seguridad)
    cols = [
      "fecha","hora","especialidad","nombres","apellidos","edad","direccion",
      "motivo_consulta","historia_enfermedad_actual",
      "app_1","app_2","app_3",
      "medicamentos_1","medicamentos_2","medicamentos_3",
      "aqx_1","aqx_2","aqx_3",
      "apnp_1","apnp_2","apnp_3",
      "alergicos",
      "obst_g","obst_p","obst_a","obst_c","obst_fum",
      "hab_alcohol","hab_cafe","hab_cafe_tazas","hab_tabaco","hab_vaper","hab_medicamentos","hab_medicamentos_cual",
      "sv_ta","sv_fc","sv_fcf","sv_fr","sv_glicemia","sv_temp","sv_sao2","sv_peso",
      "glasgow_o","glasgow_v","glasgow_m","glasgow_total",
      "insp_aspecto_general","insp_cabeza","insp_cuello","insp_torax","insp_pulmones","insp_corazon","insp_abdomen",
      "insp_genitales","insp_extremidades","insp_tacto_rectal","insp_neurologico",
      "reportes","laboratorio",
      "img_ekg","img_radiografia","img_sonografia","img_tomografia",
      "comentarios",
      "diag_1","diag_2","diag_3","diag_4","diag_5",
      "medicos",
    ]

    payload = {k: data.get(k) for k in cols}

    if not prev:
        fields = ["ingreso_id","created_at","updated_at","created_by_user","created_by_role","updated_by_user","updated_by_role"] + cols
        qs = ",".join(["?"] * len(fields))
        cur.execute(
            f"INSERT INTO ingresos_historia_clinica({','.join(fields)}) VALUES ({qs})",
            [iid, now, now, u["username"], u["role"], u["username"], u["role"]] + [payload[k] for k in cols]
        )
    else:
        set_sql = ",".join([f"{k}=?" for k in cols] + ["updated_at=?","updated_by_user=?","updated_by_role=?"])
        cur.execute(
            f"UPDATE ingresos_historia_clinica SET {set_sql} WHERE ingreso_id=?",
            [payload[k] for k in cols] + [now, u["username"], u["role"], iid]
        )

    conn.commit()
    conn.close()

    return jsonify({"ok": True})

@APP.put("/api/ingresos/<int:iid>/admision")
@require_login
def api_ingresos_update_admision(iid: int):
    u = session["user"]
    if not _can_ingresos_write_admision(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    data = request.get_json(force=True, silent=True) or {}

    allowed = {
        "paciente_master_id",
        "codigo", "fecha", "hora", "habitacion", "especialidad",
        "nombres", "apellidos", "edad", "sexo", "direccion", "cedula",
        "seguro_ars", "nss", "ocupacion",
        "fam_conyugue_acompanante", "fam_madre", "fam_padre", "fam_direccion", "fam_telefono",
        "dx_1", "dx_2", "dx_3", "dx_4", "dx_5",
    }

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM ingresos_admision WHERE ingreso_id=?", (iid,))
    prev = cur.fetchone()
    if not prev:
        cur.execute("""
          INSERT INTO ingresos_admision(ingreso_id, updated_at, updated_by_user, updated_by_role)
          VALUES (?,?,?,?)
        """, (iid, now_local(), u["username"], u["role"]))
        cur.execute("SELECT * FROM ingresos_admision WHERE ingreso_id=?", (iid,))
        prev = cur.fetchone()

    before = dict(prev)

    sets = []
    vals = []
    for k, v in data.items():
        if k in allowed:
            sets.append(f"{k}=?")
            vals.append(v)

    sets.extend(["updated_at=?", "updated_by_user=?", "updated_by_role=?"])
    vals.extend([now_local(), u["username"], u["role"]])

    cur.execute(f"UPDATE ingresos_admision SET {', '.join(sets)} WHERE ingreso_id=?", (*vals, iid))
    conn.commit()

    cur.execute("SELECT * FROM ingresos_admision WHERE ingreso_id=?", (iid,))
    after = dict(cur.fetchone())

    conn.close()

    audit("UPDATE_INGRESO_ADMISION", entity="ingresos", entity_id=iid, before=before, after=after)
    return jsonify({"ok": True})


# ---------- Evoluciones ----------
@APP.post("/api/ingresos/<int:iid>/evoluciones")
@require_login
def api_ingresos_add_evolucion(iid: int):
    u = session["user"]
    if not _can_ingresos_write_medico(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    data = request.get_json(force=True, silent=True) or {}
    fecha_hora = (data.get("fecha_hora") or now_local()).strip()
    nota = (data.get("nota") or "").strip()
    if not nota:
        return jsonify({"ok": False, "error": "nota requerida"}), 400

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM ingresos WHERE id=?", (iid,))
    if not cur.fetchone():
        conn.close()
        return jsonify({"ok": False, "error": "Ingreso no existe"}), 404

    cur.execute("""
      INSERT INTO ingreso_evoluciones(
        ingreso_id, fecha_hora, nota,
        created_at, updated_at, created_by_user, created_by_role
      ) VALUES (?,?,?,?,?,?,?)
    """, (iid, fecha_hora, nota, now_local(), now_local(), u["username"], u["role"]))
    eid = cur.lastrowid
    conn.commit()
    conn.close()

    audit("CREATE_EVOLUCION", entity="ingreso_evoluciones", entity_id=eid, after={"ingreso_id": iid, "user": u["username"]})
    return jsonify({"ok": True, "id": eid})


@APP.put("/api/ingresos/evoluciones/<int:eid>")
@require_login
def api_ingresos_update_evolucion(eid: int):
    u = session["user"]
    if not _can_ingresos_write_medico(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    data = request.get_json(force=True, silent=True) or {}
    fecha_hora = (data.get("fecha_hora") or "").strip()
    nota = (data.get("nota") or "").strip()

    if not nota:
        return jsonify({"ok": False, "error": "nota requerida"}), 400

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM ingreso_evoluciones WHERE id=?", (eid,))
    prev = cur.fetchone()
    if not prev:
        conn.close()
        return jsonify({"ok": False, "error": "No existe"}), 404

    before = dict(prev)

    cur.execute("""
      UPDATE ingreso_evoluciones
      SET fecha_hora=?, nota=?, updated_at=?
      WHERE id=?
    """, (fecha_hora or prev["fecha_hora"], nota, now_local(), eid))

    conn.commit()
    conn.close()

    audit("UPDATE_EVOLUCION", entity="ingreso_evoluciones", entity_id=eid, before=before, after={"fecha_hora": fecha_hora, "nota": nota})
    return jsonify({"ok": True})


# ---------- Órdenes médicas ----------
DEFAULT_MEDIDAS_GENERALES = (
    "1- Ingresar a Sala Clinica\n"
    "2- Signos Vitales por Turno\n"
    "3- Dieta Suave\n"
    "4- Cuidados Generales por Enfermeria\n"
    "5- Avisar Ante Eventualidad\n"
)
@APP.post("/api/ingresos/<int:iid>/ordenes")
@require_login
def api_ingresos_add_orden(iid: int):
    u = session["user"]
    if not _can_ingresos_write_medico(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    conn = db()
    cur = conn.cursor()

    data = request.get_json(force=True, silent=True) or {}
    copy_last = bool(data.get("copy_last") or False)
    fecha_hora = (data.get("fecha_hora") or now_local()).strip()

    # NUEVO: aceptar texto desde frontend
    medidas = (data.get("medidas_generales") or "").strip()
    ordenes = (data.get("ordenes") or "").strip()

    # defaults si no mandan texto
    if not medidas:
        medidas = DEFAULT_MEDIDAS_GENERALES
    if not ordenes:
        ordenes = "1- "

    # copiar última si aplica (sobrescribe)
    if copy_last:
        cur.execute("""
          SELECT medidas_generales, ordenes
          FROM ingreso_ordenes
          WHERE ingreso_id=?
          ORDER BY datetime(fecha_hora) DESC, id DESC
          LIMIT 1
        """, (iid,))
        last = cur.fetchone()
        if last:
            medidas = last["medidas_generales"] or medidas
            ordenes = last["ordenes"] or ordenes

    # validación mínima
    if not medidas or not ordenes:
        conn.close()
        return jsonify({"ok": False, "error": "Medidas generales y órdenes son requeridas."}), 400

    # INSERT (aquí es donde va)
    cur.execute("""
      INSERT INTO ingreso_ordenes(
        ingreso_id, fecha_hora, medidas_generales, ordenes,
        created_at, updated_at, created_by_user, created_by_role
      ) VALUES (?,?,?,?,?,?,?,?)
    """, (iid, fecha_hora, medidas, ordenes, now_local(), now_local(), u["username"], u["role"]))
    oid = cur.lastrowid

    conn.commit()
    conn.close()

    audit(
      "CREATE_ORDEN_MEDICA",
      entity="ingreso_ordenes",
      entity_id=oid,
      after={"ingreso_id": iid, "copy_last": copy_last}
    )
    return jsonify({"ok": True, "id": oid})

@APP.put("/api/ingresos/ordenes/<int:oid>")
@require_login
def api_ingresos_update_orden(oid: int):
    u = session["user"]
    if not _can_ingresos_write_medico(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    data = request.get_json(force=True, silent=True) or {}
    fecha_hora = (data.get("fecha_hora") or "").strip()
    medidas = (data.get("medidas_generales") or "").strip()
    ordenes = (data.get("ordenes") or "").strip()

    # valida mínimos (en edición no deberías “inventar defaults” si vienen vacíos;
    # pero si quieres mantener tu comportamiento, déjalo así)
    if not medidas:
        return jsonify({"ok": False, "error": "medidas_generales requerido"}), 400
    if not ordenes:
        return jsonify({"ok": False, "error": "ordenes requerido"}), 400

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM ingreso_ordenes WHERE id=?", (oid,))
    prev = cur.fetchone()
    if not prev:
        conn.close()
        return jsonify({"ok": False, "error": "No existe"}), 404

    before = dict(prev)

    cur.execute("""
      UPDATE ingreso_ordenes
      SET fecha_hora=?,
          medidas_generales=?,
          ordenes=?,
          updated_at=?
      WHERE id=?
    """, (
        fecha_hora or prev["fecha_hora"],
        medidas,
        ordenes,
        now_local(),
        oid
    ))

    conn.commit()
    conn.close()

    audit(
        "UPDATE_ORDEN_MEDICA",
        entity="ingreso_ordenes",
        entity_id=oid,
        before=before,
        after={
            "fecha_hora": fecha_hora or prev["fecha_hora"],
            "updated_by": u["username"]
        }
    )
    return jsonify({"ok": True})

# ---------- Epicrisis / Egreso ----------
@APP.post("/api/ingresos/<int:iid>/epicrisis")
@require_login
def api_ingresos_save_epicrisis(iid: int):
    u = session["user"]
    if not _can_ingresos_write_medico(u["role"]):
        return jsonify({"ok": False, "error": "No permitido"}), 403

    data = request.get_json(force=True, silent=True) or {}

    fecha_egreso = (data.get("fecha_egreso") or now_local()).strip()
    destino = (data.get("destino") or "").strip().lower()
    referido_hospital = (data.get("referido_hospital") or "").strip()

    allowed_dest = {"alta_medica", "referido", "fallecido", "fuga", "alta_peticion"}
    if destino not in allowed_dest:
        return jsonify({"ok": False, "error": "destino inválido"}), 400
    if destino == "referido" and not referido_hospital:
        return jsonify({"ok": False, "error": "Debes especificar hospital referido"}), 400

    payload = {
        "dx_ingreso": (data.get("dx_ingreso") or "").strip(),
        "dx_egreso": (data.get("dx_egreso") or "").strip(),
        "procedimientos": (data.get("procedimientos") or "").strip(),
        "hallazgos": (data.get("hallazgos") or "").strip(),
        "tratamiento": (data.get("tratamiento") or "").strip(),
        "plan_seguimiento": (data.get("plan_seguimiento") or "").strip(),
        "causas_fallecimiento": (data.get("causas_fallecimiento") or "").strip(),
    }

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM ingresos WHERE id=?", (iid,))
    ing_prev = cur.fetchone()
    if not ing_prev:
        conn.close()
        return jsonify({"ok": False, "error": "Ingreso no existe"}), 404

    before_ing = dict(ing_prev)

    cur.execute("SELECT * FROM ingreso_epicrisis WHERE ingreso_id=?", (iid,))
    epi_prev = cur.fetchone()
    before_epi = dict(epi_prev) if epi_prev else None

    if epi_prev:
        cur.execute("""
          UPDATE ingreso_epicrisis
          SET fecha_egreso=?, destino=?, referido_hospital=?,
              dx_ingreso=?, dx_egreso=?, procedimientos=?, hallazgos=?, tratamiento=?, plan_seguimiento=?, causas_fallecimiento=?,
              updated_at=?, created_by_user=?, created_by_role=?
          WHERE ingreso_id=?
        """, (
            fecha_egreso, destino, referido_hospital,
            payload["dx_ingreso"], payload["dx_egreso"], payload["procedimientos"], payload["hallazgos"], payload["tratamiento"], payload["plan_seguimiento"], payload["causas_fallecimiento"],
            now_local(), u["username"], u["role"],
            iid
        ))
    else:
        cur.execute("""
          INSERT INTO ingreso_epicrisis(
            ingreso_id,
            fecha_egreso, destino, referido_hospital,
            dx_ingreso, dx_egreso, procedimientos, hallazgos, tratamiento, plan_seguimiento, causas_fallecimiento,
            created_at, updated_at, created_by_user, created_by_role
          ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            iid,
            fecha_egreso, destino, referido_hospital,
            payload["dx_ingreso"], payload["dx_egreso"], payload["procedimientos"], payload["hallazgos"], payload["tratamiento"], payload["plan_seguimiento"], payload["causas_fallecimiento"],
            now_local(), now_local(), u["username"], u["role"]
        ))

    # Marcar egresado en ingreso
    cur.execute("""
      UPDATE ingresos
      SET estado='egresado',
          updated_at=?
      WHERE id=?
    """, (now_local(), iid))

    conn.commit()
    conn.close()

    audit("SAVE_EPICRISIS_EGRESO", entity="ingresos", entity_id=iid,
          before={"ingreso": before_ing, "epicrisis": before_epi},
          after={"destino": destino, "referido_hospital": referido_hospital, "fecha_egreso": fecha_egreso, **payload})
    return jsonify({"ok": True, "estado": "egresado"})


# =========================
# Boot
# =========================

if __name__ == "__main__":
    init_db()
    port = CONFIG.get("port", DEFAULT_PORT)
    APP.run(host="0.0.0.0", port=port, debug=False)